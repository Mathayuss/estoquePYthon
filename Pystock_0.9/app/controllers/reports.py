from __future__ import annotations

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.database import session_scope
from app.paths import get_exports_dir
from app.models import Product, AssetUnit, AssetMovement

class ReportController:
    @staticmethod
    def _default_path(prefix: str) -> str:
        out_dir = get_exports_dir()
        suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
        return str(out_dir / f"{prefix}_{suffix}.xlsx")

    @staticmethod
    def export_assets_xlsx(path: str|None=None) -> str:
        path = path or ReportController._default_path("ativos")
        wb = Workbook()
        ws = wb.active
        ws.title = "Ativos"

        headers = ["Patrimônio", "Serial", "Produto/Modelo", "Status", "QR", "Observações", "Criado em"]
        ws.append(headers)

        with session_scope() as s:
            rows = s.execute(
                select(AssetUnit).options(joinedload(AssetUnit.product)).order_by(AssetUnit.asset_tag.asc())
            ).scalars().all()

            for a in rows:
                ws.append([
                    a.asset_tag,
                    a.serial_number or "",
                    a.product.name if a.product else "",
                    a.status,
                    a.qr_code,
                    a.notes or "",
                    a.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                ])

        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 40

        wb.save(path)
        return path

    @staticmethod
    def export_products_xlsx(path: str|None=None) -> str:
        path = path or ReportController._default_path("produtos")
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        headers = ["Produto/Modelo", "Categoria", "Fornecedor", "Custo (R$)", "Estoque Mínimo", "Em Estoque (unid.)"]
        ws.append(headers)

        with session_scope() as s:
            prods = s.execute(
                select(Product).options(joinedload(Product.category), joinedload(Product.supplier)).order_by(Product.name.asc())
            ).scalars().all()

            for p in prods:
                count = len(s.execute(select(AssetUnit.id).where(AssetUnit.product_id == p.id, AssetUnit.status == "IN_STOCK")).all())
                ws.append([
                    p.name,
                    p.category.name if p.category else "",
                    p.supplier.name if p.supplier else "",
                    float(p.cost_price or 0.0),
                    int(p.min_stock or 0),
                    count,
                ])

        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        ws.column_dimensions["A"].width = 40

        wb.save(path)
        return path

    @staticmethod
    def export_movements_xlsx(path: str|None=None, limit: int=5000) -> str:
        path = path or ReportController._default_path("movimentacoes")
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimentacoes"

        headers = ["Data/Hora", "Tipo", "Patrimônio", "Serial", "Produto/Modelo", "Motivo", "Usuário", "Obs"]
        ws.append(headers)

        with session_scope() as s:
            rows = s.execute(
                select(AssetMovement).options(
                    joinedload(AssetMovement.asset).joinedload(AssetUnit.product),
                    joinedload(AssetMovement.user),
                    joinedload(AssetMovement.reason),
                ).order_by(AssetMovement.occurred_at.desc(), AssetMovement.id.desc()).limit(int(limit))
            ).scalars().all()

            for m in rows:
                a = m.asset
                ws.append([
                    m.occurred_at.strftime("%Y-%m-%d %H:%M:%S"),
                    m.type,
                    a.asset_tag if a else "",
                    a.serial_number if a else "",
                    a.product.name if (a and a.product) else "",
                    m.reason.name if m.reason else "",
                    m.user.username if m.user else "",
                    m.notes or "",
                ])

        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["H"].width = 40

        wb.save(path)
        return path
